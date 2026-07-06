"""
================================================================================
Match Performance Dashboard — Preprocessor
================================================================================

OVERVIEW
--------
This script reads two Excel files and produces four JSON files that power the
static dashboard (index.html).  It is intentionally flexible so it can be
adapted to any league, season, or data format — not just Allsvenskan 2026.

To adapt it to a different league or data source:
  1. Edit LEAGUE_CONFIG (league name, season, teams per league, etc.)
  2. Edit the file name / sheet name settings in FILE_CONFIG if needed.
  3. Update MATCH_FILE_LAYOUT to describe where match metadata lives in the
     2026-style file (which row = headers, which columns = team, date, etc.)
  4. Update the KPIS list — each KPI maps to a column, identified either by
     its exact header text (flexible, recommended) or by column number
     (backward-compatible with the original Allsvenskan setup).

COLUMN RESOLUTION
-----------------
For col_2026 / col_2025 you can supply:
  • An integer  — uses that exact column number (fast, order-dependent).
  • A string    — searches row 2 of the sheet for a cell whose text contains
                  the string (case-insensitive). Finds the first match.
                  Much more robust when columns shift between file versions.
  • A list      — the KPI value is the SUM of the resolved columns in the list.
  • None        — no data available from the file for this KPI.

Example — column by name:
    "Shots on Target"          # finds the first column whose header contains this

Example — column by number (original Allsvenskan mapping, still works):
    425                        # reads column 425 directly

INPUTS  (place files in source-data/)
--------------------------------------
  [reference_file]   — One row per team, full-season averages for the
                        reference/previous season.  Used to calculate rank-group
                        thresholds (1-4 / 5-8 / 9-12 / 13-16) and team averages
                        that are shown as the comparison baseline in the dashboard.

  [match_file]       — One row per match, current-season data.
                        New rows are added after each matchday.

OUTPUTS  (written to data/)
----------------------------
  config.json     — League/season metadata read by the dashboard at load time.
  kpis.json       — KPI definitions (phase, subgroup, format, etc.)
  thresholds.json — Per-KPI rank boundaries + per-team reference averages.
  matches.json    — All match records with KPI values for each team.

HOW TO RUN
----------
  pip install openpyxl        # one-time install
  python preprocess.py

================================================================================
"""

import json, re, sys
from pathlib import Path
from openpyxl import load_workbook


# ── File paths ────────────────────────────────────────────────────────────────
BASE    = Path(__file__).parent
OUT_DIR = BASE / "data"
OUT_DIR.mkdir(exist_ok=True)


# ════════════════════════════════════════════════════════════════════════════════
# LEAGUE_CONFIG  ←  Edit this section when adapting to a new league / season
# ════════════════════════════════════════════════════════════════════════════════
LEAGUE_CONFIG = {
    # Displayed in the dashboard header and browser tab title.
    "league_name":   "Allsvenskan",

    # The season being tracked (the "current" season with match-by-match data).
    "season":        "2026",

    # The reference season used for rank thresholds and baselines.
    "reference_season": "2025",

    # Total matches per team across the full season (league + cup combined).
    # Allsvenskan has 30 league games; the 4 Svenska Cupen rounds are included
    # in this file before the league starts, giving a max of 34.
    # Increase this number if more cup games are added mid-season.
    "total_matches_per_team": 34,

    # Number of matches shown in the sliding window at one time.
    "window_size": 5,

    # Number of teams in the league.
    # Used to validate that we have enough data to fill all four rank groups.
    "teams_in_league": 16,
}


# ════════════════════════════════════════════════════════════════════════════════
# FILE_CONFIG  ←  Edit if your files use different names or sheet names
# ════════════════════════════════════════════════════════════════════════════════
FILE_CONFIG = {
    # Paths to the two source Excel files.
    "reference_file": BASE / "source-data" / "Allsvenskan_2025.xlsx",
    "match_file":     BASE / "source-data" / "Allsvenskan_2026.xlsx",

    # Name of the data sheet inside each workbook.
    "reference_sheet": "data",
    "match_sheet":     "data",

    # In the reference file: which row holds KPI column headers?
    # (Row 1 is often a section/phase header; row 2 is the KPI name.)
    "reference_header_row": 2,

    # First row of team data in the reference file (rows before this are headers).
    "reference_data_start_row": 3,

    # In the match file: which row holds KPI column headers?
    "match_header_row": 2,

    # First row of match data in the match file.
    "match_data_start_row": 3,
}


# ════════════════════════════════════════════════════════════════════════════════
# MATCH_FILE_LAYOUT  ←  Describes how match metadata is stored in the match file
# ════════════════════════════════════════════════════════════════════════════════
# This tells the script where to find the team name, date, and opponent for each
# match row.  Column values can be integers (column index) or header-name strings.

MATCH_FILE_LAYOUT = {
    # Column containing the matchday label e.g. "Matchday 1".
    "matchday_col": 1,

    # Column containing the team name (and optionally H/A and score).
    # The cell is expected to look like "AIK Solna (H) 2:1"
    # or just "AIK Solna" if the format is simpler.
    "team_col": 2,

    # Column containing the match date.
    "date_col": 3,

    # Column containing the opposition team name.
    # Set to None if your file doesn't have a separate opposition column.
    "opposition_col": 4,
}


# ════════════════════════════════════════════════════════════════════════════════
# REFERENCE_FILE_LAYOUT  ←  Describes how team data is stored in the reference file
# ════════════════════════════════════════════════════════════════════════════════
REFERENCE_FILE_LAYOUT = {
    # Column containing the team name in the reference file.
    "team_col": 1,
}


# ════════════════════════════════════════════════════════════════════════════════
# TEAM_NAME_MAP  ←  Normalise team names that differ between the two files
# ════════════════════════════════════════════════════════════════════════════════
# Keys are raw names as they appear in the files; values are the canonical names
# used throughout the dashboard JSON and logo filenames.
# Add entries here whenever team names don't match exactly between files.

TEAM_NAME_MAP = {
    "AIK Solna":     "AIK",
    "GAIS Göteborg": "GAIS",
}


# ════════════════════════════════════════════════════════════════════════════════
# TEAM_TO_LOGO  ←  Maps canonical team names to logo PNG filenames in logos/
# ════════════════════════════════════════════════════════════════════════════════
# The dashboard looks for logos/<value>.png.  If a team isn't listed here, it
# will show a "?" placeholder.  Add entries when adding new teams.

TEAM_TO_LOGO = {
    "AIK":               "AIK",
    "BK Häcken":         "BK_Hacken",
    "Degerfors IF":      "Degerfors_IF",
    "Djurgårdens IF":    "Djurgardens_IF",
    "GAIS":              "GAIS",
    "Halmstads BK":      "Halmstads_BK",
    "Hammarby IF":       "Hammarby_IF",
    "IF Brommapojkarna": "IF_Brommapojkarna",
    "IF Elfsborg":       "IF_Elfsborg",
    "IFK Göteborg":      "IFK_Goteborg",
    "IK Sirius":         "IK_Sirius",
    "IK Oddevold":       "IK_Oddevold",
    "Kalmar FF":         "Kalmar_FF",
    "Malmö FF":          "Malmo_FF",
    "Mjällby AIF":       "Mjallby_AIF",
    "Västerås SK":       "Vasteras_SK",
    "Örgryte IS":        "Orgryte_IS",
}


# ════════════════════════════════════════════════════════════════════════════════
# KPIS  ←  The main list to edit when adding, removing, or changing metrics
# ════════════════════════════════════════════════════════════════════════════════
#
# Each entry is a tuple with 10 fields:
#
#   (phase_key,         — "offensive" or "defensive"
#    subgroup_key,      — "offensive_kpis", "finishing_kpis",
#                          "defensive_kpis", or "prevent_finishing"
#    phase_label,       — Tab label shown in the UI  e.g. "Offensive"
#    subgroup_label,    — Strip header shown in the grid  e.g. "Offensive KPIs"
#    kpi_name,          — Row label shown in the grid  e.g. "Touches final 3rd"
#    col_match,         — Column in the MATCH file  (int, str, list, or None)
#    col_reference,     — Column in the REFERENCE file  (int, str, list, or None)
#    higher_is_better,  — True if a higher value is better  e.g. True for shots
#    format,            — Display format:
#                            "int"   → rounded integer       e.g. 23
#                            "pct"   → percentage            e.g. 54.3%
#                            "xg"    → two decimal places    e.g. 1.23
#                            "float" → two decimal places    e.g. 10.83
#                            "goals" → one decimal/integer   e.g. 1.5
#    default_on)        — True = shown on page load; False = hidden by default
#
# ── HOW TO SPECIFY COLUMNS ────────────────────────────────────────────────────
#
# By exact column number (original Allsvenskan mapping, still works perfectly):
#   col_match = 286
#
# By header text (recommended for new leagues — robust to column shifts):
#   col_match = "Offensive Touches - Final Third"
#   The script searches for a cell in the header row whose text CONTAINS this
#   string (case-insensitive, partial match OK).
#
# As a sum of multiple columns (computed KPI):
#   col_match = [239, 250]                  # column numbers
#   col_match = ["Phase 1 xG", "Phase 2 xG"]  # header names
#   col_match = [239, "Phase 2 xG"]         # mixed (not recommended but works)
#
# Not available (no data source for this file):
#   col_match = None    # value cells show "—" without rank colour
# ─────────────────────────────────────────────────────────────────────────────

KPIS = [

    # ═══════════════════════════════════════════════════════════════════════════
    # OFFENSIVE PHASE
    # ═══════════════════════════════════════════════════════════════════════════

    # ── Offensive KPIs ────────────────────────────────────────────────────────

    # Touches recorded inside the opponent's final third of the pitch.
    # Reference 2025 col 416 = "Offensive - Thirds - Offensive Touches → Final Third"
    ("offensive", "offensive_kpis", "Offensive", "Offensive KPIs",
     "Touches final 3rd",              288,  416, True,  "int",   True),

    # Touches recorded inside the opponent's penalty box.
    # Reference 2025 col 417 = "Offensive - Thirds - Offensive Touches → Opponent Box"
    ("offensive", "offensive_kpis", "Offensive", "Offensive KPIs",
     "Touches in the box",             289,  417, True,  "int",   True),

    # Percentage of time the team is in possession of the ball.
    ("offensive", "offensive_kpis", "Offensive", "Offensive KPIs",
     "Ball Possession Rate %",         417,  545, True,  "pct",   True),

    # Forward ball progressions (passes or carries moving significantly toward goal).
    ("offensive", "offensive_kpis", "Offensive", "Offensive KPIs",
     "Ball Progression",                25,   34, True,  "int",   True),

    # Successful actions that bypassed the opponent's defensive structure.
    # Reference 2025 col 42 = Goal-threat section → Breaking Opponent Defence
    ("offensive", "offensive_kpis", "Offensive", "Offensive KPIs",
     "Breaking Opponent Defence",       32,   42, True,  "int",   True),

    # Ratio of successful offensive ball-control actions.
    ("offensive", "offensive_kpis", "Offensive", "Offensive KPIs",
     "Ratio - offensive ball control",  41,   51, True,  "pct",   True),

    # ── Finishing KPIs ────────────────────────────────────────────────────────

    # Pre-shot expected goals based on shot location and type.
    ("offensive", "finishing_kpis", "Offensive", "Finishing KPIs",
     "Shot-based xG",                   23,   21, True,  "xg",    True),

    # Post-shot xG — adjusted for where the ball actually went on target.
    ("offensive", "finishing_kpis", "Offensive", "Finishing KPIs",
     "Post-shot xG",                    24,   22, True,  "xg",    True),

    # Total shots taken.
    ("offensive", "finishing_kpis", "Offensive", "Finishing KPIs",
     "Shots",                          426,  554, True,  "int",   True),

    # Shots taken from inside the penalty box.
    # Reference 2025 col 4 = new column added in the current file version.
    ("offensive", "finishing_kpis", "Offensive", "Finishing KPIs",
     "Shots Inside the box",             9,    4, True,  "int",   True),

    # Shots that were on target (goalkeeper save or goal).
    ("offensive", "finishing_kpis", "Offensive", "Finishing KPIs",
     "Shots on Target",                427,  555, True,  "int",   True),

    # ═══════════════════════════════════════════════════════════════════════════
    # DEFENSIVE PHASE
    # ═══════════════════════════════════════════════════════════════════════════

    # ── Defensive KPIs ────────────────────────────────────────────────────────

    # Opponent touches in the team's own final third.  Lower is better.
    # Defensive - Thirds - Offensive Touches (Opponent) → First Third
    # Using (section, header) tuple so it's robust to column shifts between file versions.
    # The section "Defensive - Thirds - Offensive Touches (Opponent)" with sub-header
    # "First Third" is the correct source for this KPI.
    # Reference 2025: same section → First Third (AIK=86, col 441)
    ("defensive", "defensive_kpis", "Defensive", "Defensive KPIs",
     "Opp Touches final 3rd",
     ("Defensive - Thirds - Offensive Touches (Opponent)", "First Third"),
     ("Defensive - Thirds - Offensive Touches (Opponent)", "First Third"),
     False, "int", True),

    # Opponent touches inside the team's own penalty box.  Lower is better.
    # Defensive - Thirds - Offensive Touches (Opponent) → Own Box
    # Using (section, header) tuple so it's robust to column shifts between file versions.
    # Reference 2025: same section → Own Box (AIK=11, col 440)
    ("defensive", "defensive_kpis", "Defensive", "Defensive KPIs",
     "Opp Touches in the box",
     ("Defensive - Thirds - Offensive Touches (Opponent)", "Own Box"),
     ("Defensive - Thirds - Offensive Touches (Opponent)", "Own Box"),
     False, "int", True),

    # PPDA — Passes Allowed Per Defensive Action.
    # Lower = more aggressive pressing.  e.g. PPDA 8 = very intense press.
    ("defensive", "defensive_kpis", "Defensive", "Defensive KPIs",
     "PPDA",                            11,    2, False, "float", True),

    # Average pitch height at which the team applies pressure.  Higher = higher press.
    ("defensive", "defensive_kpis", "Defensive", "Defensive KPIs",
     "Average Pressure Height",        432,  560, True,  "float", True),

    # Times the opponent broke through the team's defensive structure.  Lower is better.
    # Col AH (34) = Own Defence Broken (summary column, updated from col 32).
    ("defensive", "defensive_kpis", "Defensive", "Defensive KPIs",
     "Own Defence Broken",              34,   27, False, "int",   True),

    # Opponent forward progressions.  Lower is better.
    # Col AB (28) = Opponent Ball Progression (summary column, updated from col 26).
    ("defensive", "defensive_kpis", "Defensive", "Defensive KPIs",
     "Opp Ball Progression",            28,   24, False, "int",   True),

    # Percentage of 1v1 duels won.
    ("defensive", "defensive_kpis", "Defensive", "Defensive KPIs",
     "Duel Rate %",                    421,  549, True,  "pct",   True),

    # Average pressure intensity in counter-pressing sequences.
    ("defensive", "defensive_kpis", "Defensive", "Defensive KPIs",
     "Avg pressure Counter Press",     464,  564, True,  "float", True),

    # Ratio of ball wins made specifically by defenders.
    ("defensive", "defensive_kpis", "Defensive", "Defensive KPIs",
     "Offensive Ball Wins (Def.)",      31,   41, True,  "pct",   True),

    # Ratio of defensive interventions relative to total defensive actions.
    ("defensive", "defensive_kpis", "Defensive", "Defensive KPIs",
     "Ratio - def interv/clearances",   37,   47, True,  "pct",   True),

    # Ratio of last-ditch interventions.  Lower is better (earlier defending).
    ("defensive", "defensive_kpis", "Defensive", "Defensive KPIs",
     "Ratio - Last Ditch Interv.",      39,   49, False, "pct",   True),

    # ── Prevent Finishing ─────────────────────────────────────────────────────
    # All metrics here are "lower is better" — fewer / lower opponent shots is good.

    # Opponent pre-shot expected goals.
    ("defensive", "prevent_finishing", "Defensive", "Prevent Finishing",
     "Opp Shot-based xG",               5,   31, False, "xg",    True),

    # Opponent post-shot xG.
    ("defensive", "prevent_finishing", "Defensive", "Prevent Finishing",
     "Opp Post-shot xG",                6,   32, False, "xg",    True),

    # Total opponent shots.
    ("defensive", "prevent_finishing", "Defensive", "Prevent Finishing",
     "Opp Shots",                       7,  402, False, "int",   True),

    # Opponent shots from inside the box.
    ("defensive", "prevent_finishing", "Defensive", "Prevent Finishing",
     "Opp Shots Inside the box",        8,    3, False, "int",   True),

    # Opponent shots on target.
    ("defensive", "prevent_finishing", "Defensive", "Prevent Finishing",
     "Opp Shots on Target",            10,    5, False, "int",   True),
]


# ════════════════════════════════════════════════════════════════════════════════
# PHASE / SUBGROUP ORDER  ←  Controls display order in the dashboard
# ════════════════════════════════════════════════════════════════════════════════

PHASE_ORDER = [
    ("offensive", "Offensive"),
    ("defensive", "Defensive"),
]

SUBGROUP_ORDER = [
    ("offensive_kpis",    "Offensive KPIs"),
    ("finishing_kpis",    "Finishing KPIs"),
    ("defensive_kpis",    "Defensive KPIs"),
    ("prevent_finishing", "Prevent Finishing"),
]


# ════════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS  — no need to edit these
# ════════════════════════════════════════════════════════════════════════════════

def kpi_key(phase, subgroup, name):
    """Build a unique string key used across all JSON files."""
    return f"{phase}::{subgroup}::{name}"


def is_computed(col):
    """True if the column value is a list (sum of multiple columns)."""
    return isinstance(col, list)


def resolve_col(ws, col_spec, header_row):
    """
    Resolve a column specification to an integer column index.

    col_spec can be:
      int          — returned as-is (column number, 1-based)
      str          — searches the header row for the first cell whose text contains
                     col_spec (case-insensitive, partial match).
      (str, str)   — tuple of (section_keyword, header_keyword).
                     Finds the first column where row-1 text contains section_keyword
                     AND row-2 text contains header_keyword (both case-insensitive).
                     Use this when the same header appears in multiple sections.
                     Example: ("Defensive - Thirds - Offensive Touches (Opponent)", "Own Box")
      list         — returns a list of resolved column indices (for computed KPIs).
      None         — returns None (no data available).

    ws         : openpyxl worksheet object
    col_spec   : the column specification from the KPIS definition
    header_row : which row holds column headers (1-based); section headers are in row header_row-1
    """
    if col_spec is None:
        return None

    if isinstance(col_spec, int):
        return col_spec

    if isinstance(col_spec, tuple) and len(col_spec) == 2:
        sec_kw = col_spec[0].lower().strip()
        hdr_kw = col_spec[1].lower().strip()
        sec_row = header_row - 1   # section labels sit one row above the header row
        last_sec = ""
        for c in range(1, ws.max_column + 1):
            sec_val = ws.cell(row=sec_row, column=c).value
            if sec_val:
                last_sec = str(sec_val).lower()
            hdr_val = ws.cell(row=header_row, column=c).value
            if hdr_val and sec_kw in last_sec and hdr_kw in str(hdr_val).lower():
                return c
        print(f"  WARNING: section='{col_spec[0]}' / header='{col_spec[1]}' not found.",
              file=sys.stderr)
        return None

    if isinstance(col_spec, str):
        search = col_spec.lower().strip()
        for c in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=header_row, column=c).value
            if cell_val and search in str(cell_val).lower():
                return c
        print(f"  WARNING: header '{col_spec}' not found — KPI will have no data.",
              file=sys.stderr)
        return None

    if isinstance(col_spec, list):
        return [resolve_col(ws, item, header_row) for item in col_spec]

    return None


def read_cell(ws, row, col_spec, header_row):
    """
    Read a numeric value from ws at (row, col_spec).
    col_spec is already resolved (int or list of ints).
    Returns float or None.
    """
    if col_spec is None:
        return None
    if isinstance(col_spec, list):
        parts = [ws.cell(row=row, column=c).value for c in col_spec if c is not None]
        valid = [v for v in parts if v is not None]
        return sum(float(v) for v in valid) if valid else None
    try:
        v = ws.cell(row=row, column=col_spec).value
        return float(v) if v is not None else None
    except (TypeError, ValueError):
        return None


def parse_team_col(raw):
    """
    Parse a combined team/result cell from the match file.

    Handles formats like:
      "AIK Solna (H) 2:1"  →  ("AIK", "H", "2:1")
      "AIK Solna"          →  ("AIK", None, None)

    Returns (team_name, home_away, score).
    home_away is "H", "A", or None.
    """
    if not raw:
        return None, None, None
    raw = str(raw).strip()
    m = re.match(r"^(.*?)\s*\(([HA])\)\s*(.*)$", raw)
    if m:
        name = TEAM_NAME_MAP.get(m.group(1).strip(), m.group(1).strip())
        return name, m.group(2), m.group(3).strip()
    return TEAM_NAME_MAP.get(raw, raw), None, None


def resolve_fixed_col(layout_value, ws, header_row):
    """
    Resolve a layout column (from MATCH_FILE_LAYOUT / REFERENCE_FILE_LAYOUT).
    Can be int or string header name.
    """
    return resolve_col(ws, layout_value, header_row)


# ════════════════════════════════════════════════════════════════════════════════
# MAIN — load files, compute thresholds, write JSON
# ════════════════════════════════════════════════════════════════════════════════

# Validate source files exist
ref_path   = FILE_CONFIG["reference_file"]
match_path = FILE_CONFIG["match_file"]
for p in (ref_path, match_path):
    if not Path(p).exists():
        print(f"ERROR: {p} not found.\n"
              f"  Place both Excel files in source-data/", file=sys.stderr)
        sys.exit(1)

# Load workbooks
wb_ref   = load_workbook(ref_path,   data_only=True)
wb_match = load_workbook(match_path, data_only=True)
ws_ref   = wb_ref[FILE_CONFIG["reference_sheet"]]
ws_match = wb_match[FILE_CONFIG["match_sheet"]]

ref_hdr_row   = FILE_CONFIG["reference_header_row"]
match_hdr_row = FILE_CONFIG["match_header_row"]
ref_start     = FILE_CONFIG["reference_data_start_row"]
match_start   = FILE_CONFIG["match_data_start_row"]

# Resolve match file layout columns
_mc = MATCH_FILE_LAYOUT
m_matchday_col  = resolve_fixed_col(_mc["matchday_col"],  ws_match, match_hdr_row)
m_team_col      = resolve_fixed_col(_mc["team_col"],      ws_match, match_hdr_row)
m_date_col      = resolve_fixed_col(_mc["date_col"],      ws_match, match_hdr_row)
m_opp_col       = (resolve_fixed_col(_mc["opposition_col"], ws_match, match_hdr_row)
                   if _mc.get("opposition_col") is not None else None)

# Resolve reference file layout column
r_team_col = resolve_fixed_col(
    REFERENCE_FILE_LAYOUT["team_col"], ws_ref, ref_hdr_row)

# Read all teams from the reference file
teams_ref = []
for r in range(ref_start, ws_ref.max_row + 1):
    name_raw = ws_ref.cell(row=r, column=r_team_col).value if r_team_col else None
    if not name_raw:
        continue
    teams_ref.append({
        "row":  r,
        "name": TEAM_NAME_MAP.get(str(name_raw).strip(), str(name_raw).strip()),
    })
print(f"Reference file: {len(teams_ref)} teams in {LEAGUE_CONFIG['reference_season']}")

# Pre-resolve all KPI columns once (avoids repeated header scans)
kpi_resolved = []  # list of (p, sg, pl, sgl, nm, resolved_c_match, resolved_c_ref, higher, fmt, default)
for p, sg, pl, sgl, nm, c_match, c_ref, higher, fmt_str, default in KPIS:
    rc_match = resolve_col(ws_match, c_match, match_hdr_row)
    rc_ref   = resolve_col(ws_ref,   c_ref,   ref_hdr_row)
    kpi_resolved.append((p, sg, pl, sgl, nm, rc_match, rc_ref, higher, fmt_str, default))


# ── Thresholds + per-team reference averages ─────────────────────────────────
thresholds = {}
team_avgs  = {}
min_teams  = 4  # need at least 4 to fill the top rank group

for p, sg, _, _, nm, rc_match, rc_ref, higher, fmt_str, default in kpi_resolved:
    key = kpi_key(p, sg, nm)
    if rc_ref is None:
        continue

    # Read every team's season average for this KPI
    vals = []
    for t in teams_ref:
        v = read_cell(ws_ref, t["row"], rc_ref, ref_hdr_row)
        if v is not None:
            vals.append((t["name"], v))

    # Store each team's value as their personal baseline
    for team_name, val in vals:
        team_avgs.setdefault(team_name, {})[key] = val

    if len(vals) < min_teams:
        continue

    # Sort best → worst and pick the three boundary values
    vals.sort(key=lambda x: x[1], reverse=higher)
    sv = [v for _, v in vals]

    thresholds[key] = {
        "higher_is_better": higher,
        "t1":  sv[3],                          # top-4 boundary
        "t2":  sv[7],                          # top-8 boundary
        "t3":  sv[min(11, len(sv)-1)],         # top-12 boundary (safe for small leagues)
        "avg": sum(sv) / len(sv),
        "min": min(sv),
        "max": max(sv),
        "ranked_teams": [
            {"team": team, "value": val, "rank": i + 1}
            for i, (team, val) in enumerate(vals)
        ],
    }


# ── Match-by-match data ───────────────────────────────────────────────────────
matches_by_team = {}

for r in range(match_start, ws_match.max_row + 1):
    # Read match metadata from their resolved columns
    matchday = (ws_match.cell(row=r, column=m_matchday_col).value
                if m_matchday_col else None)
    team_raw = (ws_match.cell(row=r, column=m_team_col).value
                if m_team_col else None)
    date_raw = (ws_match.cell(row=r, column=m_date_col).value
                if m_date_col else None)
    opp_raw  = (ws_match.cell(row=r, column=m_opp_col).value
                if m_opp_col else None)

    if not team_raw:
        continue  # skip blank rows

    team, home_away, score = parse_team_col(team_raw)
    record = {
        "matchday":  matchday,
        "opponent":  str(opp_raw).strip() if opp_raw else None,
        "home_away": home_away,
        "score":     score,
        "date":      str(date_raw).strip() if date_raw else None,
        "values":    {},
    }

    # Read KPI values for this row
    for p, sg, _, _, nm, rc_match, rc_ref, higher, fmt_str, default in kpi_resolved:
        key = kpi_key(p, sg, nm)
        if rc_match is None:
            continue
        v = read_cell(ws_match, r, rc_match, match_hdr_row)
        if v is not None:
            record["values"][key] = v

    matches_by_team.setdefault(team, []).append(record)


# ── Scan ALL named columns from the match file → columns.json ─────────────────
#
# This builds a catalogue of every column available in the match file so the
# dashboard's "+ KPI" picker can expose them all, not just the 27 pre-defined ones.
#
# For each column we store:
#   col         — column number (1-based)
#   section     — row-1 section header (e.g. "Offensive - Thirds - Offensive Touches")
#   header      — row-2 KPI header     (e.g. "Final Third")
#   label       — "Section → Header" display string (de-duplicated where needed)
#   match_values— {team_name: [val_m1, val_m2, ...]} per match order
#   ref_avg     — {team_name: avg} from the reference file (if found by column header search)

# Columns used as match metadata — skip these (they're not KPIs)
_meta_cols = {
    m_matchday_col, m_team_col, m_date_col,
    *([] if m_opp_col is None else [m_opp_col]),
}

# Columns already covered by the defined KPIS list (skip duplicates)
_defined_match_cols = set()
for _p, _sg, _pl, _sgl, _nm, _rc_m, _rc_r, *_ in kpi_resolved:
    if _rc_m is None:
        continue
    if isinstance(_rc_m, list):
        _defined_match_cols.update(c for c in _rc_m if c)
    elif _rc_m:
        _defined_match_cols.add(_rc_m)

# Build a lookup: header text → reference column index (for ref avg)
# We search the ref file header row for matching text.
_ref_header_lookup = {}
for _c in range(1, ws_ref.max_column + 1):
    _h = ws_ref.cell(row=ref_hdr_row, column=_c).value
    if _h:
        _ref_header_lookup[str(_h).strip().lower()] = _c

# Build a lookup: team name → row in the reference file
_ref_team_row = {t["name"]: t["row"] for t in teams_ref}

# Build _team_ws_rows once before the column loop: {team_name: [worksheet row numbers]}
_team_ws_rows = {}
for _r in range(match_start, ws_match.max_row + 1):
    _raw = ws_match.cell(row=_r, column=m_team_col).value if m_team_col else None
    if not _raw:
        continue
    _t, _, _ = parse_team_col(_raw)
    _team_ws_rows.setdefault(_t, []).append(_r)

# Walk every named column in the match file
all_columns = []
_last_section = ""
for _col in range(1, ws_match.max_column + 1):
    # Skip metadata columns and already-defined KPI columns
    if _col in _meta_cols or _col in _defined_match_cols:
        continue

    _section = str(ws_match.cell(row=1, column=_col).value or "").strip()
    _header  = str(ws_match.cell(row=2, column=_col).value or "").strip()
    if not _header:
        continue

    # Keep track of the last seen section header (it spans multiple columns)
    if _section:
        _last_section = _section
    _display_section = _last_section

    # Build a readable label: "Section → Header" or just "Header" if no section
    if _display_section and _display_section.lower() != _header.lower():
        _label = f"{_display_section} → {_header}"
    else:
        _label = _header

    # Collect per-team match values
    _match_values = {}
    for _team, _ws_rows in _team_ws_rows.items():
        _vals = []
        for _r in _ws_rows:
            _v = ws_match.cell(row=_r, column=_col).value
            try:
                _vals.append(float(_v) if _v is not None else None)
            except (TypeError, ValueError):
                _vals.append(None)
        if any(v is not None for v in _vals):
            _match_values[_team] = _vals

    if not _match_values:
        continue   # skip columns with no usable data

    # Look up reference average from the reference file by matching header text
    _ref_avg = {}
    _ref_col = _ref_header_lookup.get(_header.lower())
    if _ref_col:
        for _tname, _trow in _ref_team_row.items():
            _rv = ws_ref.cell(row=_trow, column=_ref_col).value
            try:
                if _rv is not None:
                    _ref_avg[_tname] = float(_rv)
            except (TypeError, ValueError):
                pass

    all_columns.append({
        "col":          _col,
        "section":      _display_section,
        "header":       _header,
        "label":        _label,
        "match_values": _match_values,
        "ref_avg":      _ref_avg,
    })

print(f"✓  Catalogued {len(all_columns)} additional columns for the picker")

# ── Build kpis.json ───────────────────────────────────────────────────────────
kpis_out = {
    "phases":    [{"key": k, "label": l} for k, l in PHASE_ORDER],
    "subgroups": [{"key": k, "label": l} for k, l in SUBGROUP_ORDER],
    "items": [
        {
            "key":              kpi_key(p, sg, nm),
            "phase":            p,
            "phase_label":      pl,
            "subgroup":         sg,
            "subgroup_label":   sgl,
            "name":             nm,
            "higher_is_better": higher,
            "format":           fmt_str,
            "default":          default,
            "manual":           rc_match is None,
            "has_reference_data": rc_ref is not None,
        }
        for p, sg, pl, sgl, nm, rc_match, rc_ref, higher, fmt_str, default in kpi_resolved
    ],
}

# ── Write all JSON output files ───────────────────────────────────────────────

# config.json — dashboard metadata (league name, season, slider range, etc.)
(OUT_DIR / "config.json").write_text(
    json.dumps({
        **LEAGUE_CONFIG,
        # Derive the dashboard title from league + season
        "dashboard_title": f"{LEAGUE_CONFIG['league_name']} {LEAGUE_CONFIG['season']}",
        # Key used throughout the dashboard for per-team baselines
        "reference_label": f"{LEAGUE_CONFIG['reference_season']} avg",
    }, ensure_ascii=False, indent=2),
    encoding="utf-8"
)

(OUT_DIR / "columns.json").write_text(
    json.dumps({"columns": all_columns}, ensure_ascii=False, indent=2),
    encoding="utf-8"
)

(OUT_DIR / "kpis.json").write_text(
    json.dumps(kpis_out, ensure_ascii=False, indent=2), encoding="utf-8")

(OUT_DIR / "thresholds.json").write_text(
    json.dumps({
        "season_source": LEAGUE_CONFIG["reference_season"],
        "kpis":          thresholds,
        "team_avgs":     team_avgs,
    }, ensure_ascii=False, indent=2), encoding="utf-8")

(OUT_DIR / "matches.json").write_text(
    json.dumps({
        "season":          LEAGUE_CONFIG["season"],
        "league":          LEAGUE_CONFIG["league_name"],
        "teams":           sorted(matches_by_team),
        "matches_by_team": matches_by_team,
        "team_to_logo":    TEAM_TO_LOGO,
    }, ensure_ascii=False, indent=2), encoding="utf-8")

# ── Summary ───────────────────────────────────────────────────────────────────
manual_n = sum(1 for row in kpi_resolved if row[5] is None)
print(f"\n✓  {LEAGUE_CONFIG['league_name']} {LEAGUE_CONFIG['season']}")
print(f"✓  {len(thresholds)} thresholds | {len(KPIS)} KPIs | {manual_n} without match-file column")
print(f"✓  Teams in match file: {sorted(matches_by_team)}")
for team, ms in sorted(matches_by_team.items()):
    print(f"   {team}: {len(ms)} matches, {len(ms[0]['values'])} auto-KPIs")
print(f"✓  JSON written to {OUT_DIR}/")
